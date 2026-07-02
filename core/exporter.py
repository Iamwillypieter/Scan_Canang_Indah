"""
Module export data ke berbagai format (Excel, CSV).
Tidak menggunakan pandas — pakai openpyxl (Excel) dan csv stdlib (CSV)
agar kompatibel dengan Python 3.14 dan menghindari DLL conflict.
"""
import csv
import os
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from core.data_processor import DailySummaryRecord
from config import EXPORT_FOLDER

logger = logging.getLogger(__name__)

# Kolom output sesuai spesifikasi
COLUMNS = [
    "No.",
    "Card ID",
    "Employee ID",
    "Name",
    "Depart.",
    "Date",
    "First IN",
    "Last OUT",
    "Terminal(First)",
    "Terminal(Last)",
    "Door(First)",
    "Door(Last)",
]


def _summary_to_row(s: DailySummaryRecord) -> list:
    """Konversi DailySummaryRecord ke list nilai untuk satu baris."""
    return [
        s.no,
        s.card_id,
        s.employee_id,
        s.name,
        s.department,
        s.date_str,
        s.first_in_str,
        s.last_out_str,
        s.terminal_first,
        s.terminal_last,
        s.door_first,
        s.door_last,
    ]


class DataExporter:
    """Export data attendance (First IN / Last OUT) ke Excel atau CSV."""

    def __init__(self, summaries: list[DailySummaryRecord]):
        self._summaries = summaries

    def to_excel(self, filepath: str = "") -> str:
        """Export ke file Excel (.xlsx) menggunakan openpyxl. Returns: filepath."""
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(EXPORT_FOLDER, f"absensi_{timestamp}.xlsx")

        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Absensi"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        header_align = Alignment(horizontal="center", vertical="center")

        # Tulis header
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Tulis data
        for row_idx, s in enumerate(self._summaries, start=2):
            row_data = _summary_to_row(s)
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column width
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = len(col_name)
            for row_idx in range(2, len(self._summaries) + 2):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(filepath)
        logger.info(f"Excel exported: {filepath} ({len(self._summaries)} rows)")
        return filepath

    def to_csv(self, filepath: str = "") -> str:
        """Export ke file CSV menggunakan stdlib csv. Returns: filepath."""
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(EXPORT_FOLDER, f"absensi_{timestamp}.csv")

        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        # UTF-8 BOM agar Excel bisa buka langsung tanpa encoding issue
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(COLUMNS)
            for s in self._summaries:
                writer.writerow(_summary_to_row(s))

        logger.info(f"CSV exported: {filepath} ({len(self._summaries)} rows)")
        return filepath
